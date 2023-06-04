import os
import openpyxl
from datetime import datetime, timedelta
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side


def non_empty_file_path(file_path, msg):
    if not file_path:
        print("Input is null or empty.")
        file_path = input(msg)
        non_empty_file_path(file_path, msg)
    else:
        return file_path


# Load the text file
text_file_path = input("Enter the path to the input text file: ")
text_file_path = non_empty_file_path(text_file_path, "Enter the path to the input text file: ")
output_file_path = input("Enter the location where output file will be generated: ")
output_file_path = non_empty_file_path(output_file_path, "Enter the location where output file will be generated: ")


def remove_file_if_exist(output_file_path):
    if os.path.exists(output_file_path):
        os.remove(output_file_path)
        print("File removed successfully.")
    else:
        print("File does not exist.")


def header_background_color():
    global fill
    light_blue = 'B7E1CE'
    fill = PatternFill(fill_type='solid', fgColor=light_blue)


def border_format():
    global border
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )


def format_cell(cellName, isMerge, row_to_apply):
    if isMerge:
        sheet.merge_cells(f'{cellName}{start_row}:{cellName}{end_row}')
        sheet[f'{cellName}{row_to_apply}'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    sheet[f'{cellName}{row_to_apply}'].fill = fill


def create_date_columns():
    start_date = datetime(int(year), int(month), 1)
    for day in range(num_days):
        current_date = start_date + timedelta(days=day)
        column_index = day + 3
        column_letter = openpyxl.utils.get_column_letter(column_index)
        sheet[column_letter + '1'] = current_date.strftime('%d-%b')
        format_cell(column_letter, False, 1)
        sheet[column_letter + '2'] = current_date.strftime('%a')[:3]
        format_cell(column_letter, False, 2)
        if current_date.strftime('%a')[:3] == 'Sat' or current_date.strftime('%a')[:3] == 'Sun':
            for i in range(3, len(name_list) + 3):
                sheet[column_letter + str(i)] = 'WO'


def write_data():
    global row
    for row in range(len(name_list)):
        sr_no = row + 1
        name = name_list[row]
        sheet['A' + str(row + 3)] = sr_no
        sheet['B' + str(row + 3)] = name


def apply_border():
    global row, cell, cell
    # Apply the border to all cells
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = border


remove_file_if_exist(output_file_path)

with open(text_file_path, 'r') as file:
    lines = file.readlines()
name_list = lines[1].strip().split(', ')
month_year = lines[2].strip()

header_background_color()

# Extract the values from the text file
# Parse the month and year from the text file
month, year = month_year.split(' ')

# Get the number of days in the specified month and year
num_days = (datetime(int(year), int(month) + 1, 1) - datetime(int(year), int(month), 1)).days

# Define the range of rows to merge and center
start_row = 1  # Starting row number
end_row = 2  # Ending row number

# Create a new workbook and sheet
output_workbook = openpyxl.Workbook()
sheet = output_workbook.active

# Write the column headers
sheet['A1'] = 'Sr. No'
sheet['B1'] = 'Name'

# Merge and center the rows
format_cell('A', True, 1)
format_cell('B', True, 1)

# Write the date columns
create_date_columns()

# Write the data rows
write_data()

border_format()
apply_border()

# Save the workbook

output_workbook.save(output_file_path)
