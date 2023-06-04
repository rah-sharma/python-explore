import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Alignment

# Load the workbook
workbook = openpyxl.load_workbook('sample.xlsx')

# Select the desired sheet
source_sheet = workbook['Sheet1']

legends = {
    'M': '07:00-16:00',
    'G': '09:00-18:00',
    'A': '14:00-23:00',
    'L': 'Leave',
    'H': 'Holiday',
    'WO': 'Weekly Off'
}

# Create a new sheet to copy the data
new_sheet = workbook.create_sheet('NewSheet')

# Iterate over the cells in the source sheet
def get_shift_value(cell_val):
    if cell_val == 'M':
        return legends[cell_val]
    elif cell_val == 'G':
        return legends[cell_val]
    elif cell_val == 'A':
        return legends[cell_val]
    elif cell_val == 'L':
        return legends[cell_val]
    elif cell_val == 'H':
        return legends[cell_val]
    elif cell_val == 'WO':
        return legends[cell_val]
    else:
        return cell_val


for row in source_sheet.iter_rows():
    for cell in row:
        # Get the value and formatting of the cell

        value = get_shift_value(cell.value)

        # Get the formatting of the cell
        font = cell.font
        fill = cell.fill
        border = cell.border
        alignment = cell.alignment

        # Create the same cell in the new sheet
        new_cell = new_sheet.cell(row=cell.row, column=cell.column, value=value)

        # Apply formatting to the new cell
        new_cell.font = Font(**font.__dict__)
        new_cell.fill = PatternFill(**fill.__dict__)
        new_cell.border = Border(**border.__dict__)
        new_cell.alignment = Alignment(**alignment.__dict__)

# Save the workbook
workbook.save('sample.xlsx')
